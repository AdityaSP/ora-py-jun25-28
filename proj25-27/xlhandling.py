import openpyxl
wb = openpyxl.Workbook()
sh = wb.create_sheet('My Data')
sh['A1'] = 'Hello There'
sh['A2'] = 'Created by a program'
wb.save('trial2.xlsx')


# import openpyxl
# wb = openpyxl.load_workbook('trial.xlsx')
# print wb.sheetnames
# sh = wb['Sheet1']
#
# for col in sh.rows:
#     for cell in col:
#         print cell.value
# print "*" * 50
# for row in sh.columns:
#     for cell in row:
#         print cell.value
#
# wb.close()
